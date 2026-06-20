import urllib.request
import io
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

URL = 'http://127.0.0.1:5000/user-management/bulk-import/template'

def download_template():
    try:
        resp = urllib.request.urlopen(URL)
        if resp.getcode() != 200:
            print('Failed to download template, status', resp.getcode())
            sys.exit(1)
        data = resp.read()
        return data
    except Exception as e:
        print('Error downloading template:', e)
        sys.exit(1)

def extract_ids(data):
    wb = load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb.active
    # Read header row to map column names to letters
    headers = [cell.value for cell in ws[1]]
    col_map = {header: get_column_letter(idx+1) for idx, header in enumerate(headers)}
    ids = {}
    for dv in ws.data_validations.dataValidation:
        # dv.ranges is a MultiCellRange; get first cell reference
        rng = next(iter(dv.ranges.ranges))
        # Example range: 'D2:D1048576'
        col_letter = rng.split('2')[0]  # get column letter before row numbers
        # Find which header corresponds to this column
        header = None
        for h, col in col_map.items():
            if col == col_letter:
                header = h
                break
        if header:
            # dv.formula1 is a string like "'1,2,3'"
            formula = dv.formula1.strip('"')
            # Remove surrounding single quotes if present
            if formula.startswith("'") and formula.endswith("'"):
                formula = formula[1:-1]
            values = formula.split(',') if formula else []
            ids[header] = values
    return ids

if __name__ == '__main__':
    binary = download_template()
    extracted = extract_ids(binary)
    for header, values in extracted.items():
        print(f"{header}: {', '.join(values)}")
