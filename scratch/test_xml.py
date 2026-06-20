import io
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from xml.etree import ElementTree

wb = Workbook()
ws = wb.active
ws.title = "Main"
ws_dropdown = wb.create_sheet(title="DropdownData")

ws_dropdown.cell(row=1, column=1, value="item1")
ws_dropdown.cell(row=2, column=1, value="item2")

# Case 1: Omitted
dv1 = DataValidation(type="list", formula1="DropdownData!$A$1:$A$2")
ws.add_data_validation(dv1)
dv1.ranges.add("B2:B10")

# Case 2: showDropDown=False
dv2 = DataValidation(type="list", formula1="DropdownData!$A$1:$A$2", showDropDown=False)
ws.add_data_validation(dv2)
dv2.ranges.add("C2:C10")

# Case 3: showDropDown=True
dv3 = DataValidation(type="list", formula1="DropdownData!$A$1:$A$2", showDropDown=True)
ws.add_data_validation(dv3)
dv3.ranges.add("D2:D10")

# Save workbook to memory
stream = io.BytesIO()
wb.save(stream)
data = stream.getvalue()

# Let's inspect worksheet XML in the zip file
import zipfile
with zipfile.ZipFile(io.BytesIO(data)) as z:
    for f in z.namelist():
        if "sheet1" in f:
            xml_content = z.read(f)
            print("Sheet 1 XML DataValidations:")
            root = ElementTree.fromstring(xml_content)
            ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            dvs = root.findall(".//ns:dataValidation", ns)
            for i, d in enumerate(dvs, 1):
                print(f"DV {i}:", ElementTree.tostring(d, encoding="utf-8").decode("utf-8"))
