import openpyxl
import os

wb = openpyxl.load_workbook('../VIL_project  BRD.xlsx')
output_path = 'scratch/brd_summary.md'

with open(output_path, 'w', encoding='utf-8') as f:
    f.write("# VIL Project BRD Summary\n\n")
    
    for sheet_name in wb.sheetnames:
        f.write(f"## Sheet: {sheet_name}\n\n")
        f.write("```text\n")
        sheet = wb[sheet_name]
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                # Filter trailing Nones
                last_idx = len(row)
                while last_idx > 0 and row[last_idx-1] is None:
                    last_idx -= 1
                row_filtered = row[:last_idx]
                
                if row_filtered:
                    # Format cells nicely
                    row_str = " | ".join(str(cell) if cell is not None else "" for cell in row_filtered)
                    f.write(f"Row {row_count+1}: {row_str}\n")
                    row_count += 1
        f.write("```\n\n")

print(f"BRD summary successfully written to {output_path}")
