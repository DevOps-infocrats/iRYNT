import io
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# Import the Flask app and service (adjust import paths if needed)
from app import create_app
from app.modules.users.services.excel_parser_service import ExcelParserService

app = create_app()
app.app_context().push()

service = ExcelParserService()
# Generate the template (bytes)
template_bytes = service.generate_template()

# Load the template to read dropdown validation values
wb_template = load_workbook(filename=io.BytesIO(template_bytes), data_only=True)
ws_template = wb_template.active

# After building header_to_values, write them to a text file
output_txt_path = "ids.txt"
with open(output_txt_path, "w", encoding="utf-8") as f:
    for hdr, vals in header_to_values.items():
        f.write(f"{hdr}: {', '.join(vals)}\n")
print(f"✅ IDs exported to {output_txt_path}")

    # Get first cell range (e.g., CellRange('D2:D1048576'))
    cell_range = next(iter(dv.ranges.ranges))
    # Extract column letter before the row number (e.g., 'D')
    col_letter = cell_range.coord.split('2')[0]

    # Find the header that matches this column
    header = None
    for cell in ws_template[1]:
        if get_column_letter(cell.column) == col_letter:
            header = cell.value
            break
    if header:
        # dv.formula1 is something like "'1,2,3'"
        formula = dv.formula1.strip('"')
        if formula.startswith("'") and formula.endswith("'"):
            formula = formula[1:-1]
        values = formula.split(',') if formula else []
        header_to_values[header] = values

# Create a new workbook for the test import
wb_test = Workbook()
ws_test = wb_test.active
ws_test.title = "User Import Test"

# Write the header row (same as in the template)
headers = ExcelParserService.HEADERS
for col_idx, header in enumerate(headers, start=1):
    ws_test.cell(row=1, column=col_idx, value=header)

# Populate a single row with the first available ID from each dropdown
row_vals = []
for header in headers:
    vals = header_to_values.get(header)
    if not vals:
        # For free‑text fields just use placeholder data
        if header == "Username":
            row_vals.append("test_user")
        elif header == "Email":
            row_vals.append("test_user@example.com")
        elif header == "Phone":
            row_vals.append("+1234567890")
        elif header == "Password":
            row_vals.append("Password123!")
        else:
            row_vals.append("")
    else:
        # Use the first allowed value (e.g., first Company ID, Yes/No, etc.)
        row_vals.append(vals[0])

# Write the data row
for col_idx, val in enumerate(row_vals, start=1):
    ws_test.cell(row=2, column=col_idx, value=val)

# Save the test file
output_path = "test_import.xlsx"
wb_test.save(output_path)
print(f"Test import file created: {output_path}")
print("Values used for each column:")
for hdr, val in zip(headers, row_vals):
    print(f"  {hdr}: {val}")
