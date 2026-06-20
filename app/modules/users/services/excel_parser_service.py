import io
import sys
import os

def _import_openpyxl():
    """Import required openpyxl symbols, ensuring virtualenv site-packages are on sys.path.

    Returns:
        tuple: (Workbook, load_workbook, Font, Alignment, PatternFill,
               DataValidation, get_column_letter)
    Raises:
        RuntimeError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        return Workbook, load_workbook, Font, Alignment, PatternFill, DataValidation, get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel import/export. Install it via: pip install openpyxl") from exc


class ExcelParserService:
    """Service to generate template, parse uploaded Excel, and generate error reports.

    Expected columns (in order):
        Username, Email, Phone, Role, Company, Circle, Project, Is Active, Is Verified, Password (optional)
    """

    HEADERS = [
        "Username",
        "Email",
        "Phone",
        "Role",
        "Company",
        "Circle",
        "Project",
        "Is Active",
        "Is Verified",
        "Password",
    ]

    def generate_template(self) -> bytes:
        """Generate an Excel template with dropdowns for related IDs and boolean fields.

        The dropdowns are populated from the current database contents for
        Company, Circle, and Role identifiers. ``Is Active`` and ``Is Verified``
        use a static ``Yes/No`` list.
        """
        Workbook, _, Font, Alignment, PatternFill, DataValidation, get_column_letter = _import_openpyxl()
        # Import here to avoid circular import issues at module load time.
        from app.extensions import db
        from app.modules.companies.models import Company
        from app.modules.circles.models import Circle
        from app.modules.auth.models import Role
        from app.modules.projects.models import Project

        wb = Workbook()
        ws = wb.active
        ws.title = "User Import Template"
        
        # Header row with styling
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            
        # Freeze header
        ws.freeze_panes = ws["A2"]
        
        # ---------- Data validation for dropdowns ----------
        # Fetch active names from database
        roles = [r.name for r in db.session.query(Role.name).order_by(Role.name).all()]
        companies = [c.company_name for c in db.session.query(Company.company_name).filter_by(status='Active').order_by(Company.company_name).all()]
        circles = [c.circle_name for c in db.session.query(Circle.circle_name).filter_by(status='Active').order_by(Circle.circle_name).all()]
        projects = [p.project_name for p in db.session.query(Project.project_name).filter_by(status='Active').order_by(Project.project_name).all()]
        yes_no = ["Yes", "No"]

        # Create hidden dropdown sheet
        ws_dropdown = wb.create_sheet(title="DropdownData")
        ws_dropdown.sheet_state = 'hidden'

        # Headers for DropdownData sheet
        ws_dropdown.cell(row=1, column=1, value="Roles")
        ws_dropdown.cell(row=1, column=2, value="Companies")
        ws_dropdown.cell(row=1, column=3, value="Circles")
        ws_dropdown.cell(row=1, column=4, value="Projects")
        ws_dropdown.cell(row=1, column=5, value="YesNo")

        # Populate lists
        for idx, val in enumerate(roles, start=2):
            ws_dropdown.cell(row=idx, column=1, value=val)
        for idx, val in enumerate(companies, start=2):
            ws_dropdown.cell(row=idx, column=2, value=val)
        for idx, val in enumerate(circles, start=2):
            ws_dropdown.cell(row=idx, column=3, value=val)
        for idx, val in enumerate(projects, start=2):
            ws_dropdown.cell(row=idx, column=4, value=val)
        for idx, val in enumerate(yes_no, start=2):
            ws_dropdown.cell(row=idx, column=5, value=val)

        header_to_col = {header: get_column_letter(idx) for idx, header in enumerate(self.HEADERS, start=1)}

        def apply_validation_range(target_col_letter, list_col_idx, num_items):
            if num_items == 0:
                return
            list_col_letter = get_column_letter(list_col_idx)
            formula = f"DropdownData!${list_col_letter}$2:${list_col_letter}${num_items + 1}"
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            dv.ranges.add(f"{target_col_letter}2:{target_col_letter}1048576")

        apply_validation_range(header_to_col["Role"], 1, len(roles))
        apply_validation_range(header_to_col["Company"], 2, len(companies))
        apply_validation_range(header_to_col["Circle"], 3, len(circles))
        apply_validation_range(header_to_col["Project"], 4, len(projects))
        apply_validation_range(header_to_col["Is Active"], 5, len(yes_no))
        apply_validation_range(header_to_col["Is Verified"], 5, len(yes_no))
        
        # Save to bytes
        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def parse_excel(self, uploaded_file) -> list:
        import sys, os
        print('DEBUG parse_excel: sys.executable =', sys.executable)
        print('DEBUG parse_excel: sys.path =', sys.path)
        print('DEBUG parse_excel: cwd =', os.getcwd())
        _, load_workbook, _, _, _, _, _ = _import_openpyxl()
        """Parse an uploaded Excel file (FileStorage) and return list of dict rows.
        Each row dict uses header names as keys.
        """
        wb = load_workbook(filename=uploaded_file, data_only=True)
        ws = wb.active
        rows = []
        # Assume first row is header
        header = [str(cell.value or "").strip() for cell in ws[1]]
        
        # Ensure we only process if headers are present
        if not header or not any(h in self.HEADERS for h in header):
            header = self.HEADERS
            
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # If the row is completely empty, skip it
            if not any(cell is not None for cell in row):
                continue
            row_dict = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
            row_dict["row_number"] = row_idx  # 1-based Excel row number
            rows.append(row_dict)
        return rows

    def generate_error_report(self, errors: list) -> bytes:
        import sys, os
        print('DEBUG generate_error_report: sys.executable =', sys.executable)
        print('DEBUG generate_error_report: sys.path =', sys.path)
        print('DEBUG generate_error_report: cwd =', os.getcwd())
        Workbook, _, Font, Alignment, PatternFill, _, get_column_letter = _import_openpyxl()
        """Create an Excel workbook listing all validation errors in a detailed table."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Errors"
        
        headers = ["Row Number", "Column", "Value", "Error", "Suggested Fix"]
        
        # Header row with styling
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            
        # Write error records
        fill = PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid")
        for err_idx, err in enumerate(errors, start=2):
            row_num = err.get("row_number")
            column = err.get("column", "")
            val = str(err.get("value") if err.get("value") is not None else "")
            msg = err.get("message", "")
            fix = err.get("suggested_fix", "")
            
            row_vals = [row_num, column, val, msg, fix]
            ws.append(row_vals)
            
            # Highlight the error row
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=err_idx, column=col_idx).fill = fill
                
        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()
