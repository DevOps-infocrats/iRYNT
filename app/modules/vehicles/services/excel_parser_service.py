import io
import sys
import os
from datetime import date

def _import_openpyxl():
    """Import required openpyxl symbols dynamically."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        return Workbook, load_workbook, Font, Alignment, PatternFill, DataValidation, get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel import/export. Install it via: pip install openpyxl") from exc


class VehicleExcelParserService:
    """Service to generate templates, parse Excel files, and download errors for Vehicles."""

    HEADERS = [
        "Vehicle Number",
        "Vehicle Type",
        "Vehicle Category",
        "Vehicle Brand",
        "Vehicle Model",
        "Company",
        "Circle",
        "Client",
        "Project",
        "Subzone",
        "Manufacturing Year",
        "Chassis Number",
        "Engine Number",
        "Owner Name",
        "Owner Phone",
        "Vendor Name",
        "Vendor Contact",
        "GPS Enabled",
        "Realtime Tracking Enabled",
        "Deployment Allowed"
    ]

    def generate_template(self) -> bytes:
        """Generate an Excel template with dropdown validations for vehicles."""
        Workbook, _, Font, Alignment, PatternFill, DataValidation, get_column_letter = _import_openpyxl()
        from app.extensions import db
        from app.modules.companies.models import Company
        from app.modules.circles.models import Circle
        from app.modules.clients.models import Client
        from app.modules.projects.models import Project
        from app.modules.subzones.models import Subzone
        from app.modules.vehicles.vehicle_status import VEHICLE_TYPES, VEHICLE_CATEGORIES

        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicle Import Template"

        # Style headers
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = ws["A2"]

        # Fetch active records from DB
        companies = [c.company_name for c in db.session.query(Company.company_name).filter_by(status='Active').order_by(Company.company_name).all()]
        circles = [c.circle_name for c in db.session.query(Circle.circle_name).filter_by(status='Active').order_by(Circle.circle_name).all()]
        clients = [c.client_name for c in db.session.query(Client.client_name).filter_by(status='Active').order_by(Client.client_name).all()]
        projects = [p.project_name for p in db.session.query(Project.project_name).filter_by(status='Active').order_by(Project.project_name).all()]
        subzones = [s.subzone_name for s in db.session.query(Subzone.subzone_name).filter_by(status='Active').order_by(Subzone.subzone_name).all()]
        
        # Static options
        yes_no = ["Yes", "No"]
        vehicle_types = [t[1] for t in VEHICLE_TYPES if t[0]]
        vehicle_categories = [c[1] for c in VEHICLE_CATEGORIES if c[0]]
        
        current_year = date.today().year
        mfg_years = [str(y) for y in range(current_year, current_year - 25, -1)]

        # Hidden worksheet for dropdown options
        ws_dropdown = wb.create_sheet(title="DropdownData")
        ws_dropdown.sheet_state = 'hidden'

        # Headers for DropdownData
        dropdown_cols = [
            ("Companies", companies),
            ("Circles", circles),
            ("Clients", clients),
            ("Projects", projects),
            ("Subzones", subzones),
            ("YesNo", yes_no),
            ("VehicleTypes", vehicle_types),
            ("VehicleCategories", vehicle_categories),
            ("MfgYears", mfg_years)
        ]

        for col_idx, (title, items) in enumerate(dropdown_cols, start=1):
            ws_dropdown.cell(row=1, column=col_idx, value=title)
            for row_idx, val in enumerate(items, start=2):
                ws_dropdown.cell(row=row_idx, column=col_idx, value=val)

        header_to_col = {header: get_column_letter(idx) for idx, header in enumerate(self.HEADERS, start=1)}

        def apply_validation(target_col, list_col_idx, count):
            if count == 0:
                return
            col_letter = get_column_letter(list_col_idx)
            formula = f"DropdownData!${col_letter}$2:${col_letter}${count + 1}"
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            dv.ranges.add(f"{target_col}2:{target_col}1048576")

        # Map validations to main worksheet columns
        apply_validation(header_to_col["Company"], 1, len(companies))
        apply_validation(header_to_col["Circle"], 2, len(circles))
        apply_validation(header_to_col["Client"], 3, len(clients))
        apply_validation(header_to_col["Project"], 4, len(projects))
        apply_validation(header_to_col["Subzone"], 5, len(subzones))
        apply_validation(header_to_col["GPS Enabled"], 6, len(yes_no))
        apply_validation(header_to_col["Realtime Tracking Enabled"], 6, len(yes_no))
        apply_validation(header_to_col["Deployment Allowed"], 6, len(yes_no))
        apply_validation(header_to_col["Vehicle Type"], 7, len(vehicle_types))
        apply_validation(header_to_col["Vehicle Category"], 8, len(vehicle_categories))
        apply_validation(header_to_col["Manufacturing Year"], 9, len(mfg_years))

        # Auto-adjust column widths on main sheet
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def parse_excel(self, uploaded_file) -> list:
        """Parse uploaded Excel file and return rows as dictionary list."""
        _, load_workbook, _, _, _, _, _ = _import_openpyxl()
        wb = load_workbook(filename=uploaded_file, data_only=True)
        ws = wb.active
        rows = []
        
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        if not headers or not any(h in self.HEADERS for h in headers):
            headers = self.HEADERS
            
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):
                continue
            row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            row_dict["row_number"] = row_idx
            rows.append(row_dict)
            
        return rows

    def generate_error_report(self, errors: list) -> bytes:
        """Generate a 5-column error list spreadsheet."""
        Workbook, _, Font, Alignment, PatternFill, _, get_column_letter = _import_openpyxl()
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Errors"

        headers = ["Row Number", "Column", "Value", "Error", "Suggested Fix"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        fill = PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid")
        for err_idx, err in enumerate(errors, start=2):
            row_vals = [
                err.get("row_number"),
                err.get("column", ""),
                str(err.get("value") if err.get("value") is not None else ""),
                err.get("message", ""),
                err.get("suggested_fix", "")
            ]
            ws.append(row_vals)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=err_idx, column=col_idx).fill = fill

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()
