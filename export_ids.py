import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Import the service that builds the bulk‑import template.
# No Flask app context is required; the service is self‑contained.
# Import Flask app creation and push context
from app import create_app
from app.modules.users.services.excel_parser_service import ExcelParserService
app = create_app()
app.app_context().push()
# Generate the template bytes using the service
service = ExcelParserService()
template_bytes = service.generate_template()

# ----------------------------------------------------------------------
# 2. Load the workbook and collect dropdown values for each column
# ----------------------------------------------------------------------
wb = load_workbook(filename=io.BytesIO(template_bytes), data_only=True)
ws = wb.active

header_to_values = {}

for dv in ws.data_validations.dataValidation:
    # dv.ranges.ranges is a list of CellRange objects; take the first one.
    cell_range = next(iter(dv.ranges.ranges))  # first range in the set
    col_letter = cell_range.coord.split('2')[0]

    # Find the header that lives in this column (first row).
    header = None
    for cell in ws[1]:
        if get_column_letter(cell.column) == col_letter:
            header = cell.value
            break

    if header:
        # dv.formula1 looks like "'1,2,3'" (quotes included)
        formula = dv.formula1.strip('"')
        if formula.startswith("'") and formula.endswith("'"):
            formula = formula[1:-1]
        values = formula.split(',') if formula else []
        header_to_values[header] = values

# ----------------------------------------------------------------------
# 3. Write all IDs to a simple text file (ids.txt)
# ----------------------------------------------------------------------
output_path = "ids.txt"
with open(output_path, "w", encoding="utf-8") as f:
    for hdr, vals in header_to_values.items():
        f.write(f"{hdr}: {', '.join(vals)}\n")
print(f"✅ IDs exported to {output_path}")

# Optional: pretty‑print to console for quick verification
print("\nExtracted IDs:")
for hdr, vals in header_to_values.items():
    print(f"{hdr}: {', '.join(vals)}")
